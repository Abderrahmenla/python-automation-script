import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import rows_from_range
import time
import openai
import logging
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import pyfiglet
from colorama import init, Fore, Style
import jpype
import asposecells
from asposecells.api import Workbook

# Initialize colorama for colored output in the terminal
init(autoreset=True)

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s: %(message)s')
logger = logging.getLogger(__name__)

openai.api_key = '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'


def check_string_category(prompt):
    if "EXCELLENT" in prompt[:70]:
        return 0
    elif "VERY GOOD" in prompt[:70]:
        return 1
    elif "GOOD" in prompt[:70]:
        return 2
    elif "INADEQUATE" in prompt[:70]:
        return 4
    elif "ADEQUATE" in prompt[:70]:
        return 3
    elif "NO SUBMISSION" in prompt[:70]:
        return 5    
    else:
        return -1  # Return -1 if none of the specified substrings are found

def send_prompt_to_chat(prompt):
    try:
        
        response = openai.ChatCompletion.create(
            model='gpt-3.5-turbo',
            messages=[
                {
                    'role' : 'user',
                    'content' : prompt
                }
            ] ,
            max_tokens=120,
            n=1,
            stop=None,
            temperature=0.7,
            top_p=1.0,
            frequency_penalty=0.0,
            presence_penalty=0.0
        )
        generated_text = response['choices'][0]['message']['content']
        return generated_text
    except Exception as e:
        print(Fore.RED + f"Error generating chat response: {e}")
        return ""

def get_merged_cell_value(sheet, cell):
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return cell.value

def parse_excel_sheet(sheet,start_row,end_row,start_column,end_column):
    try:
        exclude_rows = [9, 10, 19, 20, 26, 27, 32, 33, 37,38, 43, 44]
        data = []
        processed_rows = 0
        row_num=start_row
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_column, max_col=end_column):
            if row_num in exclude_rows:
                data.append([])
                row_num+=1
                continue
            merged_row_content = ""  # List to store the merged cells in each row
            print('this is the row',len(row))
            for cell in row:
                value = get_merged_cell_value(sheet, cell)
                merged_row_content = str(value)
                break
            data.append(merged_row_content) 
            processed_rows += 1
            row_num+=1

        return data
    except Exception as e:
        print(Fore.RED  + f"Error parsing Excel sheet: {e}")
        return []

def generate_prompts(array1, array2):
    if len(array1) != len(array2):
        raise ValueError("Both input arrays must have the same length.")

    merged_result = []
    for i in range(len(array1)):
        if not (len(array1[i]) and len(array2[i])):
            merged_result.append('')
            continue
        prompt = f"A student got a grade with this description: {array1[i]} Write a feedback on how he can improve and get the next grade which is: {array2[i]} in 120 words maximum" 
        merged_result.append(prompt)
    return merged_result

def fill_excel_with_data(file_path, sheet_name, data, start_column,start_row):
    try:
        wb = load_workbook(filename=file_path)
        sheet = wb[sheet_name]

        for row_index, row in enumerate(data):
            if not len(data[row_index]):
                continue
            feedback_index = check_string_category(data[row_index])
            column = start_column + feedback_index
            cell = sheet.cell(row=start_row+row_index, column=column)
            if not feedback_index:
                cell.value = "Your attainment for this criterion has been classified as “Excellent” (the maximum classification that can be achieved). All requirements for this criterion have been achieved."
                wb.save(file_path)
                print(Fore.GREEN + "Data saved to", file_path)
                continue
            generated_feedback = send_prompt_to_chat(data[row_index])
            print('\n\n\n',generated_feedback)
            time.sleep(10)
            cell.value = generated_feedback
            wb.save(file_path)
            print('---------------------------------------')
            print('---------------------------------------')
            print('---------------------------------------\n')
            print(Fore.GREEN + "Data saved to", file_path)
       
    except Exception as e:
        print(Fore.RED + f"Error filling Excel with data: {e}")

def export_excel_to_pdf(excel_file, sheet_name, pdf_file):
      jpype.startJVM() 
      workbook = Workbook(excel_file)
      workbook.save("Output.pdf")
      jpype.shutdownJVM()


if __name__ == "__main__":
    if len(sys.argv) != 4:
        print(Fore.RED + "Usage: python script.py <file_path> <start_column> <start_row>")
        sys.exit(1)

    ASCII_art = pyfiglet.figlet_format("TOM SCRIPT")
    print(ASCII_art)
    file_path = sys.argv[1]
    sheet_name = "AI Seed"  # Specify the desired sheet name here
    fill_sheet_name = "Criteria Classifications"
    start_column = ord(sys.argv[2].upper()) - 64
    start_row = int(sys.argv[3])
    try:
        workbook = load_workbook(file_path,data_only=True,keep_vba=True)
        sheet = workbook[sheet_name]
        aimed_result = parse_excel_sheet(sheet,4,50,ord('B') - 64,ord('G') - 64)
        current_result = parse_excel_sheet(sheet,4,50,ord('I') - 64,ord('I') - 64)
        generated_prompts = generate_prompts(current_result,aimed_result)
        print("generated promps : \n",generated_prompts)
        print("-------------------------------------------------------------------------------------------------------------------------------------------------------------")
        print("-------------------------------------------------------------------------------------------------------------------------------------------------------------")
        print("-------------------------------------------------------------------------------------------------------------------------------------------------------------")
        print("-------------------------------------------------------------------------------------------------------------------------------------------------------------")

        fill_excel_with_data(file_path, fill_sheet_name, generated_prompts, start_column, start_row)
        export_excel_to_pdf("prototype_v02.xlsx", "AI Seed", "grades.pdf")

        print(Fore.GREEN + "Data saved to", file_path)
        print(Fore.GREEN + "Excel to PDF conversion completed.")
    except Exception as e:
        logger.exception("An error occurred:")
        print(Fore.RED + "An error occurred. Please check the logs for more details.")

